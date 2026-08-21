import streamlit as st
import openpyxl
import pypdf
import zipfile
import os
import gc
import re
import io
import pytesseract
import pandas as pd
from io import BytesIO
from pyzbar.pyzbar import decode
from pdf2image import convert_from_bytes
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- CẤU HÌNH GIAO DIỆN STREAMLIT ---
st.set_page_config(page_title="POD Tool - Hisense & BSH", page_icon="⚡", layout="wide")
st.markdown("""
    <style>
    .reportview-container .main .block-container{ padding-top: 2rem; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #FF4B4B; color: white; font-weight: bold; }
    .stButton>button:hover { background-color: #FF3333; color: white; }
    </style>
""", unsafe_allow_html=True)

# --- THANH MENU BÊN TRÁI ---
st.sidebar.title("⚡ HỆ THỐNG XỬ LÝ POD")
st.sidebar.markdown("Vui lòng chọn tính năng cần dùng bên dưới:")

option = st.sidebar.radio(
    "CHỌN CÔNG CỤ:",
    [
        "1. Quét Mã Vạch & Tách PDF (Zip + Excel)",
        "2. BSH_OCR & Tách PDF (Zip + Excel)"
    ]
)  

# --- QUẢN LÝ PHIÊN LÀM VIỆC ---
if "current_option" not in st.session_state:
    st.session_state.current_option = option

if st.session_state.current_option != option:
    st.session_state.current_option = option
    st.session_state.zip_path = None
    st.session_state.processed_option = None
    st.session_state.opt1_records = []
    st.session_state.opt2_records = []

if "zip_path" not in st.session_state:
    st.session_state.zip_path = None

if "processed_option" not in st.session_state:
    st.session_state.processed_option = None

if "opt1_records" not in st.session_state:
    st.session_state.opt1_records = []

if "opt2_records" not in st.session_state:
    st.session_state.opt2_records = []


# --- HÀM TẠO CHỮ CÁI TỰ ĐỘNG CHO TRANG KHÔNG ĐỌC ĐƯỢC ---
def get_letter_code(index):
    result = ""
    while index >= 0:
        result = chr(65 + (index % 26)) + result
        index = (index // 26) - 1
    return result

def sanitize_filename(name):
    clean_name = re.sub(r'[\\/*?:"<>|]', '_', str(name).strip())
    clean_name = clean_name.replace(' ', '_')
    return clean_name if clean_name else "FILE_UNNAMED"

# --- HÀM TẠO ZIP RA ĐĨA CỨNG CHO OPTION 1 ---
def build_opt1_zip_and_excel(records):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh Sach POD"
 
        headers = ["STT", "STT Trang", "Tên File / Mã Vạch", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for idx, rec in enumerate(records, 1):
            row_num = idx + 1
            ws.cell(row=row_num, column=1, value=rec["stt"])
            ws.cell(row=row_num, column=2, value=rec["stt_trang"])
            ws.cell(row=row_num, column=3, value=rec["code"])
            ws.cell(row=row_num, column=4, value=rec["status"])

            if rec["status"] == "KHÔNG_ĐỌC_ĐƯỢC":
                ws.cell(row=row_num, column=4).font = openpyxl.styles.Font(color="FF0000", bold=True)
            elif rec["status"] == "UPDATED":
                ws.cell(row=row_num, column=4).font = openpyxl.styles.Font(color="0000FF", bold=True)

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()

        zip_filename = "KET_QUA_POD.zip"
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
            zipf.writestr("KET_QUA_POD/DANH_SACH_TONG_HOP_MA_VACH.xlsx", excel_bytes)
            for rec in records:
                if rec.get("pdf_bytes") and len(rec["pdf_bytes"]) > 100:
                    clean_c = sanitize_filename(rec['code'])
                    pdf_name = f"KET_QUA_POD/{clean_c}.pdf"
                    zipf.writestr(pdf_name, rec["pdf_bytes"])

        gc.collect()
        if os.path.exists(zip_filename) and os.path.getsize(zip_filename) > 500:
            return zip_filename
        return None
    except Exception as e:
        print(f"Lỗi build zip opt1: {e}")
        return None

# --- HÀM TẠO ZIP RA ĐĨA CỨNG CHO OPTION 2 ---
def build_opt2_zip_and_excel(records):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DANH_SACH_TONG_HOP"

        headers = ["STT Tổng", "File Gốc", "STT Trang", "Mã Trích Xuất", "Loại Chứng Từ", "Trạng Thái"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        zip_filename = "KET_QUA_BSH_TONG_HOP.zip"
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
            code_tracker = {}
            
            for idx, rec in enumerate(records):
                row_num = idx + 2
                ws.cell(row=row_num, column=1, value=rec["stt"])
                ws.cell(row=row_num, column=2, value=rec["orig_file"])
                ws.cell(row=row_num, column=3, value=rec["stt_trang"])
                ws.cell(row=row_num, column=4, value=rec["code"])
                ws.cell(row=row_num, column=5, value=rec["doc_type"])
                ws.cell(row=row_num, column=6, value=rec["status"])
                if rec["status"] == "KHÔNG_ĐỌC_ĐƯỢC":
                    ws.cell(row=row_num, column=4).font = openpyxl.styles.Font(color="FF0000", bold=True)
                elif rec["status"] == "UPDATED":
                    ws.cell(row=row_num, column=4).font = openpyxl.styles.Font(color="0000FF", bold=True)

                if rec.get("pdf_bytes") and len(rec["pdf_bytes"]) > 100:
                    clean_code = sanitize_filename(rec["code"])
                    code_tracker[clean_code] = code_tracker.get(clean_code, 0) + 1
                    if code_tracker[clean_code] > 1:
                        out_filename = f"{clean_code}_p{code_tracker[clean_code]}.pdf"
                    else:
                        out_filename = f"{clean_code}.pdf"

                    out_zip_path = f"PDF_DA_TACH/{out_filename}"
                    zipf.writestr(out_zip_path, rec["pdf_bytes"])

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 18

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            zipf.writestr("DANH_SACH_MA_BSH_TONG_HOP.xlsx", excel_bytes)

        gc.collect()
        if os.path.exists(zip_filename) and os.path.getsize(zip_filename) > 500:
            return zip_filename
        return None
    except Exception as e:
        print(f"Lỗi build zip opt2: {e}")
        return None


# ==========================================================================================
# OPTION 1: QUÉT MÃ VẠCH & TÁCH PDF TỰ ĐỘNG
# ==========================================================================================
if option == "1. Quét Mã Vạch & Tách PDF (Zip + Excel)":
    st.header("⚡ CÔNG CỤ QUÉT MÃ VẠCH & TÁCH PDF TỰ ĐỘNG")
    
    uploaded_pdfs = st.file_uploader("Tải các file PDF cần tách lên đây:", type=["pdf"], accept_multiple_files=True, key="opt1_pdfs")
    
    if uploaded_pdfs:
        st.info(f"📁 Đã nhận {len(uploaded_pdfs)} file PDF sẵn sàng xử lý.")
        
        if st.button("🚀 BẮT ĐẦU XỬ LÝ", key="btn_start_opt1"):
            gc.collect()
            st.session_state.zip_path = None
            st.session_state.opt1_records = []
            st.session_state.processed_option = None
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            def process_opt1_page(args):
                page_idx, pdf_bytes_inner = args
                page_num = page_idx + 1
                
                try:
                    page_images = convert_from_bytes(pdf_bytes_inner, dpi=150, first_page=page_num, last_page=page_num, thread_count=1)
                except Exception:
                    page_images = []
                
                barcode_list = []
                preview_img_bytes = None
                
                if page_images:
                    page_image = page_images[0]
                    try:
                        barcodes = decode(page_image)
                        for barcode in barcodes:
                            barcode_data = barcode.data.decode("utf-8").strip()
                            if barcode_data:
                                clean_code = "".join(c for c in barcode_data if c.isalnum() or c in (' ', '_', '-')).strip()
                                if clean_code:
                                    barcode_list.append(clean_code)
                    except Exception:
                        pass

                    img_byte_arr = BytesIO()
                    page_image.save(img_byte_arr, format='PNG')
                    preview_img_bytes = img_byte_arr.getvalue()
                    del page_image, page_images
                    gc.collect()

                try:
                    reader = pypdf.PdfReader(BytesIO(pdf_bytes_inner))
                    pdf_writer = pypdf.PdfWriter()
                    pdf_writer.add_page(reader.pages[page_idx])
                    page_buffer = BytesIO()
                    pdf_writer.write(page_buffer)
                    single_pdf_bytes = page_buffer.getvalue()
                except Exception:
                    single_pdf_bytes = b""

                return page_idx, barcode_list, preview_img_bytes, single_pdf_bytes

            try:
                all_records = []
                global_stt = 1
                unreadable_counter = 0
                total_files = len(uploaded_pdfs)
                
                for file_idx, uploaded_pdf in enumerate(uploaded_pdfs):
                    orig_filename = uploaded_pdf.name
                    status_text.text(f"⚡ Đang xử lý file ({file_idx+1}/{total_files}): {orig_filename}...")
                    
                    pdf_bytes_input = uploaded_pdf.read()
                    pdf_reader = pypdf.PdfReader(BytesIO(pdf_bytes_input))
                    total_pages = len(pdf_reader.pages)
                    
                    tasks = [(i, pdf_bytes_input) for i in range(total_pages)]
                    max_workers = min(2, os.cpu_count() or 2)
                    
                    page_results = [None] * total_pages
                    completed_count = 0
                    
                    with ThreadPoolExecutor(max_workers=max_workers) as executor:
                        future_to_task = {executor.submit(process_opt1_page, task): task[0] for task in tasks}
                        for future in as_completed(future_to_task):
                            p_idx, barcode_list, preview_img_bytes, single_pdf_bytes = future.result()
                            page_results[p_idx] = (barcode_list, preview_img_bytes, single_pdf_bytes)
                            completed_count += 1
                            progress_bar.progress(min((file_idx + (completed_count / total_pages)) / total_files, 1.0))

                    for p_idx, (barcode_list, preview_img_bytes, single_pdf_bytes) in enumerate(page_results):
                        if barcode_list:
                            code_name = "_".join(barcode_list)
                            status_str = "THÀNH_CÔNG"
                        else:
                            code_name = get_letter_code(unreadable_counter)
                            unreadable_counter += 1
                            status_str = "KHÔNG_ĐỌC_ĐƯỢC"

                        all_records.append({
                            "stt": global_stt,
                            "stt_trang": p_idx + 1,
                            "code": code_name,
                            "status": status_str,
                            "pdf_bytes": single_pdf_bytes,
                            "preview_img": preview_img_bytes,
                            "orig_file": orig_filename
                        })
                        global_stt += 1

                st.session_state.opt1_records = all_records
                zip_filepath = build_opt1_zip_and_excel(all_records)
                
                if zip_filepath:
                    st.session_state.zip_path = zip_filepath
                    st.session_state.processed_option = "opt1"
                    status_text.text("🎉 Hoàn thành xử lý!")
                    st.success(f"🎉 Đã hoàn thành quét {total_files} file PDF ({len(all_records)} trang)!")
                else:
                    st.session_state.zip_path = None
                    st.session_state.processed_option = None
                    st.error("❌ Lỗi: File ZIP tạo ra bị rỗng hoặc lỗi dữ liệu.")

            except Exception as ex:
                st.session_state.zip_path = None
                st.session_state.processed_option = None
                st.error(f"❌ Lỗi hệ thống: {str(ex)}")

    if st.session_state.processed_option == "opt1" and st.session_state.opt1_records:
        changes_to_apply = {}
        unreadable_items = [r for r in st.session_state.opt1_records if r["status"] == "KHÔNG_ĐỌC_ĐƯỢC"]
        
        if unreadable_items:
            st.write("---")
            st.warning(f"⚠️ **BƯỚC 1: XỬ LÝ {len(unreadable_items)} TRANG KHÔNG ĐỌC ĐƯỢC MÃ**")
            for item in unreadable_items:
                col1, col2 = st.columns([1, 1])
                with col1:
                    st.markdown(f"**STT {item['stt']} | File: `{item['orig_file']}` (Trang {item['stt_trang']})**")
                    st.markdown(f"Mã tạm: `:red[{item['code']}]`")
                    if item.get("preview_img"):
                        st.image(item["preview_img"], width='stretch')
                with col2:
                    st.write("")
                    new_val = st.text_input(f"Nhập Mã DO mới cho STT {item['stt']}:", key=f"bulk_input_{item['stt']}")
                    if new_val.strip():
                        changes_to_apply[item['stt']] = new_val.strip()
                st.markdown("---")

        st.write("---")
        st.subheader("🔍 BƯỚC 2: KIỂM TRA & SỬA TRANG THEO STT")
        col_s1, col_s2 = st.columns([1, 2])
        with col_s1:
            stt_to_check = st.number_input("Nhập STT trang muốn kiểm tra:", min_value=1, max_value=len(st.session_state.opt1_records), value=1, step=1)
            selected_item = next((r for r in st.session_state.opt1_records if r["stt"] == stt_to_check), None)
            if selected_item:
                st.write(f"- File: `{selected_item['orig_file']}` (Trang {selected_item['stt_trang']})")
                st.write(f"- Mã hiện tại: `{selected_item['code']}`")
                st.write(f"- Trạng thái: `{selected_item['status']}`")
                single_val = st.text_input(f"Nhập mã thay thế cho STT {stt_to_check}:", key=f"single_input_{stt_to_check}")
                if single_val.strip():
                    changes_to_apply[stt_to_check] = single_val.strip()
        with col_s2:
            if selected_item and selected_item.get("preview_img"):
                st.image(selected_item["preview_img"], width='stretch')

        st.write("---")
        if st.button("🔄 CẬP NHẬT DỮ LIỆU ĐÃ SỬA", key="btn_update_opt1"):
            if changes_to_apply:
                for rec in st.session_state.opt1_records:
                    if rec["stt"] in changes_to_apply:
                        rec["code"] = changes_to_apply[rec["stt"]]
                        rec["status"] = "UPDATED"
                new_zip = build_opt1_zip_and_excel(st.session_state.opt1_records)
                if new_zip:
                    st.session_state.zip_path = new_zip
                    st.success("✅ Cập nhật thành công!")
                    st.rerun()
                else:
                    st.error("❌ Lỗi tạo lại file ZIP.")
            else:
                st.info("💡 Chưa có mã nào được nhập.")

        st.write("---")
        st.subheader("📊 BƯỚC 3: PREVIEW EXCEL")
        df_preview = pd.DataFrame([{ "STT": r["stt"], "File": r["orig_file"], "Trang": r["stt_trang"], "Mã": r["code"], "Status": r["status"] } for r in st.session_state.opt1_records])
        st.dataframe(df_preview, width='stretch', height=250)

        st.write("---")
        st.subheader("📥 BƯỚC 4: TẢI KẾT QUẢ")
        if st.session_state.zip_path and os.path.exists(st.session_state.zip_path):
            with open(st.session_state.zip_path, "rb") as f_dl:
                st.download_button(
                    label="📥 TẢI FILE ZIP KẾT QUẢ", 
                    data=f_dl, 
                    file_name="KET_QUA_POD.zip", 
                    mime="application/zip", 
                    type="primary", 
                    key="dl_opt1"
                )

# ==========================================================================================
# OPTION 2: BSH_OCR & TÁCH PDF TỰ ĐỘNG
# ==========================================================================================
elif option == "2. BSH_OCR & Tách PDF (Zip + Excel)":
    st.header("⚡ BSH - CÔNG CỤ OCR & TÁCH PDF TỰ ĐỘNG")
    
    uploaded_files = st.file_uploader(
        "Tải các file PDF cần tách lên đây (Có thể chọn nhiều file):", 
        type=["pdf"], 
        accept_multiple_files=True, 
        key="opt2_pdf"
    )

    if uploaded_files:
        st.info(f"📁 Đã nhận **{len(uploaded_files)}** file PDF sẵn sàng xử lý.")
        
        if st.button("🚀 BẮT ĐẦU XỬ LÝ", key="btn_start_opt2"):
            gc.collect()
            st.session_state.zip_path = None
            st.session_state.opt2_records = []
            st.session_state.processed_option = None
            
            progress_bar = st.progress(0)
            status_text = st.empty()

            def process_single_page(args):
                i, pdf_bytes_inner = args
                page_num = i + 1
                
                os.environ["OMP_THREAD_LIMIT"] = "1"
                
                final_code = None
                doc_type = "PHIEU_GIAO_HANG"
                full_text = ""
                preview_img_bytes = None

                try:
                    reader_page = pypdf.PdfReader(io.BytesIO(pdf_bytes_inner))
                    full_text = reader_page.pages[i].extract_text() or ""
                    flat_text = re.sub(r'\s+', ' ', full_text).strip()
                    if re.search(r'DANH\s*SÁCH\s*TRẢ\s*HÀNG|Return\s*Delivery', flat_text, re.IGNORECASE):
                        doc_type = "DANH_SACH_TRA_HANG"
                except Exception:
                    pass

                try:
                    if full_text:
                        matches_p7 = re.findall(r'\b(7623\d{6})\b', flat_text)
                        matches_p0 = re.findall(r'\b(7620\d{6})\b', flat_text)
                        if matches_p7:
                            final_code = matches_p7[0]
                            doc_type = "DANH_SACH_TRA_HANG"
                        elif matches_p0:
                            final_code = matches_p0[0]
                            doc_type = "PHIEU_GIAO_HANG"
                except Exception:
                    pass

                try:
                    page_images = convert_from_bytes(
                        pdf_bytes_inner,
                        dpi=150,
                        first_page=page_num,
                        last_page=page_num,
                        thread_count=1
                    )
                    if page_images:
                        img = page_images[0]
                        width, height = img.size

                        img_byte_arr = BytesIO()
                        img.save(img_byte_arr, format='PNG')
                        preview_img_bytes = img_byte_arr.getvalue()

                        if not final_code:
                            box_roi = (int(width * 0.40), int(height * 0.03), width, int(height * 0.18))
                            crop_img = img.crop(box_roi).convert('L')
                            threshold = 160
                            crop_binary = crop_img.point(lambda p: 255 if p > threshold else 0)

                            ocr_text = pytesseract.image_to_string(crop_binary, lang='eng', config='--psm 7')
                            
                            if doc_type == "DANH_SACH_TRA_HANG" or "DANH SÁCH TRẢ HÀNG" in full_text.upper():
                                doc_type = "DANH_SACH_TRA_HANG"
                                matches = re.findall(r'\b(7623\d{6})\b', ocr_text)
                                if not matches:
                                    matches = re.findall(r'\b(7620\d{6})\b', ocr_text)
                            else:
                                matches = re.findall(r'\b(7620\d{6})\b', ocr_text)
                                if not matches:
                                    matches = re.findall(r'\b(7623\d{6})\b', ocr_text)

                            if matches:
                                final_code = matches[0]

                            if not final_code:
                                ocr_text_alt = pytesseract.image_to_string(crop_img, lang='eng', config='--psm 6')
                                matches_alt = re.findall(r'\b(7623\d{6}|7620\d{6})\b', ocr_text_alt)
                                if matches_alt:
                                    final_code = matches_alt[0]
                                    if matches_alt[0].startswith('7623'):
                                        doc_type = "DANH_SACH_TRA_HANG"

                            if not final_code:
                                box_wide = (0, 0, width, int(height * 0.40))
                                wide_crop = img.crop(box_wide).convert('L')
                                for th in [140, 160, 180]:
                                    wide_bin = wide_crop.point(lambda p: 255 if p > th else 0)
                                    ocr_wide = pytesseract.image_to_string(wide_bin, lang='vie+eng', config='--psm 6')
                                    
                                    matches_wide = re.findall(r'\b(7623\d{6}|7620\d{6})\b', ocr_wide)
                                    if matches_wide:
                                        final_code = matches_wide[0]
                                        if matches_wide[0].startswith('7623'):
                                            doc_type = "DANH_SACH_TRA_HANG"
                                        break

                            if not final_code:
                                ocr_full = pytesseract.image_to_string(img.convert('L'), lang='vie+eng', config='--psm 6')
                                matches_full = re.findall(r'\b(7623\d{6}|7620\d{6})\b', ocr_full)
                                if matches_full:
                                    final_code = matches_full[0]
                                    if matches_full[0].startswith('7623'):
                                        doc_type = "DANH_SACH_TRA_HANG"

                        del page_images, img
                        gc.collect()
                except Exception:
                    pass

                pdf_writer = pypdf.PdfWriter()
                reader_writer = pypdf.PdfReader(io.BytesIO(pdf_bytes_inner))
                pdf_writer.add_page(reader_writer.pages[i])

                mem_file = io.BytesIO()
                pdf_writer.write(mem_file)
                pdf_bytes_page = mem_file.getvalue()
                mem_file.close()

                status_str = "THÀNH_CÔNG" if final_code else "KHÔNG_ĐỌC_ĐƯỢC"

                return i, page_num, final_code, doc_type, status_str, preview_img_bytes, pdf_bytes_page

            try:
                all_records = []
                global_stt = 1
                unreadable_counter = 0
                total_files = len(uploaded_files)

                for file_idx, uploaded_file in enumerate(uploaded_files):
                    orig_filename = uploaded_file.name

                    status_text.text(f"🔄 Đang xử lý file ({file_idx+1}/{total_files}): {orig_filename}...")

                    pdf_bytes = uploaded_file.read()
                    reader_temp = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                    total_pages = len(reader_temp.pages)

                    tasks = [(i, pdf_bytes) for i in range(total_pages)]
                    max_workers = min(2, os.cpu_count() or 2)

                    page_results = [None] * total_pages
                    completed_count = 0
                    with ThreadPoolExecutor(max_workers=max_workers) as executor:
                        future_to_task = {executor.submit(process_single_page, task): task[0] for task in tasks}
                        for future in as_completed(future_to_task):
                            i, page_num, final_code, doc_type, status_str, preview_img_bytes, pdf_bytes_page = future.result()
                            page_results[i] = (page_num, final_code, doc_type, status_str, preview_img_bytes, pdf_bytes_page)
                            completed_count += 1
                            overall_progress = (file_idx + (completed_count / total_pages)) / total_files
                            progress_bar.progress(min(overall_progress, 1.0))

                    for p_idx, (page_num, final_code, doc_type, status_str, preview_img_bytes, pdf_bytes_page) in enumerate(page_results):
                        if not final_code:
                            final_code = get_letter_code(unreadable_counter)
                            unreadable_counter += 1

                        all_records.append({
                            "stt": global_stt,
                            "stt_trang": page_num,
                            "code": final_code,
                            "doc_type": doc_type,
                            "status": status_str,
                            "preview_img": preview_img_bytes,
                            "pdf_bytes": pdf_bytes_page,
                            "orig_file": orig_filename
                        })
                        global_stt += 1

                st.session_state.opt2_records = all_records
                st.session_state.zip_path = build_opt2_zip_and_excel(all_records)
                st.session_state.processed_option = "opt2"
                
                status_text.text("🎉 Hoàn tất toàn bộ các file!")
                st.success(f"🎉 Đã OCR & xử lý thành công {total_files} file PDF ({len(all_records)} trang).")

            except Exception as ex:
                st.error(f"❌ Lỗi hệ thống: {str(ex)}")

    if st.session_state.get("processed_option") == "opt2" and st.session_state.get("opt2_records"):
        changes_to_apply = {}

        unreadable_items = [r for r in st.session_state.opt2_records if r["status"] == "KHÔNG_ĐỌC_ĐƯỢC"]
        if unreadable_items:
            st.write("---")
            st.warning(f"⚠️ **BƯỚC 1: XỬ LÝ {len(unreadable_items)} TRANG KHÔNG ĐỌC ĐƯỢC MÃ (BSH OCR)**")
            
            for item in unreadable_items:
                col1, col2 = st.columns([1, 1])
                with col1:
                    st.markdown(f"**STT {item['stt']} | File gốc: `{item['orig_file']}` (Trang {item['stt_trang']})**")
                    st.markdown(f"Loại chứng từ: `{item['doc_type']}`")
                    st.markdown(f"Mã tạm: `:red[{item['code']}]`")
                    if item.get("preview_img"):
                        st.image(item["preview_img"], caption=f"Hình ảnh trang {item['stt_trang']}", width='stretch')
                
                with col2:
                    st.write("")
                    st.write("")
                    new_val = st.text_input(
                        f"Nhập mã mới cho STT {item['stt']}:", 
                        key=f"opt2_bulk_input_{item['stt']}",
                        placeholder="Ví dụ: 7620123456..."
                    )
                    if new_val.strip():
                        changes_to_apply[item['stt']] = new_val.strip()
                st.markdown("---")

        st.write("---")
        st.subheader("🔍 BƯỚC 2: KIỂM TRA & SỬA TRANG THEO STT TỔNG HỢP")
        
        col_search1, col_search2 = st.columns([1, 2])
        with col_search1:
            stt_to_check = st.number_input(
                "Nhập STT trang muốn kiểm tra/sửa lại:", 
                min_value=1, 
                max_value=len(st.session_state.opt2_records), 
                value=1,
                step=1
            )
            
            selected_item = next((r for r in st.session_state.opt2_records if r["stt"] == stt_to_check), None)
            
            if selected_item:
                st.markdown(f"**Đang chọn STT {stt_to_check}:**")
                st.write(f"- File gốc: `{selected_item['orig_file']}` (Trang {selected_item['stt_trang']})")
                st.write(f"- Loại chứng từ: `{selected_item['doc_type']}`")
                st.write(f"- Mã trích xuất: `{selected_item['code']}`")
                st.write(f"- Trạng thái: `{selected_item['status']}`")
                
                single_val = st.text_input(
                    f"Nhập mã thay thế cho STT {stt_to_check}:", 
                    key=f"opt2_single_input_{stt_to_check}",
                    placeholder="Nhập mã mới..."
                )
                if single_val.strip():
                    changes_to_apply[stt_to_check] = single_val.strip()

        with col_search2:
            if selected_item and selected_item.get("preview_img"):
                st.image(
                    selected_item["preview_img"], 
                    caption=f"Hình ảnh trang {selected_item['stt_trang']} của file {selected_item['orig_file']}", 
                    width='stretch'
                )

        st.write("---")
        if st.button("🔄 BẤM VÀO ĐÂY ĐỂ CẬP NHẬT DỮ LIỆU ĐÃ SỬA", key="btn_update_opt2"):
            if changes_to_apply:
                for rec in st.session_state.opt2_records:
                    if rec["stt"] in changes_to_apply:
                        new_code = changes_to_apply[rec["stt"]]
                        rec["code"] = new_code
                        rec["status"] = "UPDATED"
                        
                        if new_code.startswith("7623"):
                            rec["doc_type"] = "DANH_SACH_TRA_HANG"
                        elif new_code.startswith("7620"):
                            rec["doc_type"] = "PHIEU_GIAO_HANG"
                
                st.session_state.zip_path = build_opt2_zip_and_excel(st.session_state.opt2_records)
                st.success("✅ Đã cập nhật dữ liệu BSH thành công! Bảng Preview đã làm mới.")
                st.rerun()
            else:
                st.info("💡 Bạn chưa nhập mã mới nào ở Bước 1 hoặc Bước 2.")

        st.write("---")
        st.subheader("📊 BƯỚC 3: PREVIEW BẢNG EXCEL TỔNG HỢP BSH (LIVE UPDATE)")
        
        preview_data = []
        for r in st.session_state.opt2_records:
            preview_data.append({
                "STT Tổng": r["stt"],
                "File Gốc": r["orig_file"],
                "STT Trang": r["stt_trang"],
                "Mã Trích Xuất": r["code"],
                "Loại Chứng Từ": r["doc_type"],
                "Trạng Thái": r["status"]
            })
        
        df_preview = pd.DataFrame(preview_data)
        st.dataframe(df_preview, width='stretch', height=300)

        st.write("---")
        st.subheader("📥 BƯỚC 4: TẢI FILE KẾT QUẢ KHI ĐÃ OK")
        
        if st.session_state.zip_path and os.path.exists(st.session_state.zip_path):
            with open(st.session_state.zip_path, "rb") as f_dl2:
                st.download_button(
                    label="📥 BẤM VÀO ĐÂY ĐỂ TẢI (ZIP TỔNG HỢP BSH)",
                    data=f_dl2,
                    file_name="KET_QUA_BSH_TONG_HOP.zip",
                    mime="application/zip",
                    type="primary",
                    key="download_opt2_zip_final"
                )
